from __future__ import annotations
import streamlit as st
import pandas as pd
import re
import sqlite3
from pathlib import Path
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="PUF Bundle Plan Generator", layout="wide")
st.title("PUF Bundle Plan Generator")
st.caption("Upload Planning File + Job Card Export -> Auto-generates Bundle Plan")

DB_PATH = Path(__file__).resolve().parent / "bundle_plan.db"
DEFAULT_CREDS_PATH = str(Path(__file__).resolve().parent / "service_account.json")

LAYER_DATA = {
    "30MM WALL PANEL": {"normal_size": 36, "threshold": 10000, "large_size": 10},
    "40MM WALL PANEL": {"normal_size": 26, "threshold": None,  "large_size": None},
    "50MM WALL PANEL": {"normal_size": 22, "threshold": 10000, "large_size": 10},
    "60MM WALL PANEL": {"normal_size": 18, "threshold": None,  "large_size": None},
    "80MM WALL PANEL": {"normal_size": 13, "threshold": None,  "large_size": None},
    "100MM WALL PANEL":{"normal_size": 11, "threshold": None,  "large_size": None},
    "120MM WALL PANEL":{"normal_size": 9,  "threshold": None,  "large_size": None},
    "150MM WALL PANEL":{"normal_size": 7,  "threshold": None,  "large_size": None},
    "20MM ROOF PANEL": {"normal_size": 26, "threshold": None,  "large_size": None},
    "30MM ROOF PANEL": {"normal_size": 22, "threshold": 8000,  "large_size": 12},
    "40MM ROOF PANEL": {"normal_size": 18, "threshold": None,  "large_size": None},
    "50MM ROOF PANEL": {"normal_size": 16, "threshold": 8000,  "large_size": 10},
    "60MM ROOF PANEL": {"normal_size": 14, "threshold": None,  "large_size": None},
    "80MM ROOF PANEL": {"normal_size": 10, "threshold": None,  "large_size": None},
    "100MM ROOF PANEL":{"normal_size": 8,  "threshold": None,  "large_size": None},
    "120MM ROOF PANEL":{"normal_size": 8,  "threshold": None,  "large_size": None},
}

@dataclass
class PanelGroup:
    length: float
    quantity: int

@dataclass
class Bundle:
    bundle_no: int
    panels: list = field(default_factory=list)

    @property
    def total_panels(self):
        return sum(pg.quantity for pg in self.panels)

    @property
    def composition(self):
        parts = []
        for i, pg in enumerate(self.panels, start=1):
            l = str(int(pg.length)) if pg.length == int(pg.length) else str(pg.length)
            parts.append(f"{pg.quantity} x L{i}({l})")
        return " + ".join(parts)

def _distribute_prefer_even(total, n, bundle_size):
    """Distribute total into n bundle sizes, minimizing odd-sized bundles.

    For roof panels, even-numbered bundles are easier to store and stack.
    - If total is even: achieves 0 odd bundles
    - If total is odd: achieves exactly 1 odd bundle (minimum possible)
    Falls back to standard distribution if optimization would create
    invalid sizes (<=0 or exceeding bundle_size).
    """
    if n <= 0:
        return []
    if n == 1:
        return [total]

    base = total // n
    remainder = total % n

    # Fallback: standard distribution
    standard = [total // n + (1 if i < total % n else 0) for i in range(n)]

    # Guard: if base is too small, redistribution could create 0-size bundles
    if base < 2:
        return standard

    if base % 2 == 0:
        # base is even, base+1 is odd
        # Standard gives `remainder` odd bundles (those that get base+1)
        # Fix: redistribute +1s in pairs of +2, leaving at most one +1
        if remainder == 0:
            return [base] * n  # already all even
        sizes = [base] * n
        pairs = remainder // 2
        single = remainder % 2
        for i in range(pairs):
            sizes[i] += 2  # base+2 is still even
        if single:
            sizes[n - 1] += 1  # one unavoidable odd bundle, placed last
    else:
        # base is odd, base+1 is even
        # Standard gives `n - remainder` odd bundles (those that get just base)
        # Fix: pair up odd bundles → one gets base-1 (even), other gets base+1 (even)
        sizes = [base + 1] * remainder  # already even
        n_odd = n - remainder
        pairs = n_odd // 2
        single = n_odd % 2
        for _ in range(pairs):
            sizes.append(base + 1)  # even
            sizes.append(base - 1)  # even
        if single:
            sizes.append(base)  # one unavoidable odd

    # Safety: ensure no size is <=0 or exceeds bundle_size
    if any(s <= 0 for s in sizes) or any(s > bundle_size for s in sizes):
        return standard

    # Sort: even bundles first (descending), then odd bundles (descending)
    # This ensures longer panels go into even-sized bundles
    sizes.sort(key=lambda x: (x % 2, -x))

    return sizes


def create_bundles(panel_groups, bundle_size, prefer_even=False):
    total = sum(pg.quantity for pg in panel_groups)
    if total == 0:
        return []
    n = max(1, (total + bundle_size - 1) // bundle_size)
    if prefer_even and n > 1:
        sizes = _distribute_prefer_even(total, n, bundle_size)
    else:
        sizes = [total // n + (1 if i < total % n else 0) for i in range(n)]
    sorted_groups = sorted(panel_groups, key=lambda x: x.length, reverse=True)
    pool = [[pg.length, pg.quantity] for pg in sorted_groups]
    bundles = []
    bundle_no = 1
    for target in sizes:
        current = Bundle(bundle_no=bundle_no)
        remaining_target = target
        for item in pool:
            if remaining_target <= 0:
                break
            if item[1] <= 0:
                continue
            take = min(item[1], remaining_target)
            if take > 0:
                current.panels.append(PanelGroup(length=item[0], quantity=take))
                item[1] -= take
                remaining_target -= take
        bundles.append(current)
        bundle_no += 1
    return bundles

def extract_colour(val):
    if pd.isna(val):
        return "UNKNOWN"
    s = str(val).strip()
    cleaned = re.sub(r"^[0-9]+\.?[0-9]*\s*", "", s).strip()
    return cleaned.upper() if cleaned else s.upper()

def parse_material(material_val):
    if pd.isna(material_val):
        return None, None
    d = str(material_val).strip().upper()
    ptype = "R" if "ROOF" in d else ("W" if "WALL" in d else None)
    m = re.search(r"(\d+)\s*MM", d)
    thickness = int(m.group(1)) if m else None
    return ptype, thickness

def parse_description(desc):
    if pd.isna(desc):
        return None, None
    d = str(desc).strip().upper()
    if "ROOF" in d or "TILE" in d or "SOLAR" in d:
        ptype = "R"
    elif "WALL" in d:
        ptype = "W"
    else:
        return None, None
    m = re.search(r"THICK[:\s]+(\d+)", d)
    thickness = int(m.group(1)) if m else None
    return ptype, thickness


def generate_label_html(bundle_df: pd.DataFrame) -> str:
    def parse_comp(comp):
        items = re.findall(r'(\d+)\s*x\s*L\d+\(([^)]+)\)', str(comp))
        out = []
        for qty, ln in items:
            try:
                out.append((str(int(float(ln))), str(int(float(qty)))))
            except Exception:
                out.append((str(ln), str(qty)))
        total = sum(int(q) for _, q in out)
        return out, total

    cards = []
    for work_id, grp in bundle_df.groupby("WORK ID", sort=False):
        grp = grp.reset_index(drop=True)
        total_bundles = len(grp)
        for i, row in grp.iterrows():
            lengths, total_qty = parse_comp(row["COMPOSITION"])
            w = row["WIDTH (m)"]
            width_disp = (str(w).rstrip("0").rstrip(".") + " M") if pd.notna(w) else ""
            region = str(row.get("REGION", "")).strip() if pd.notna(row.get("REGION", "")) and str(row.get("REGION", "")).strip() else "--"
            cards.append({
                "bundle_id": str(row["BUNDLE ID"]),
                "so": str(row["S.O"]),
                "type": str(row["TYPE"]),
                "thickness": str(row["THICKNESS"]),
                "colour": str(row["COLOUR"]),
                "width": width_disp,
                "region": region,
                "current": f"{i+1:02d}",
                "total": f"{total_bundles:02d}",
                "lengths": lengths,
                "total_qty": total_qty,
            })

    def build_table(length_pairs, total_qty=None, include_total=False):
        len_cells = "".join([f"<td class='len'>{ln}</td>" for ln, _ in length_pairs])
        qty_cells = "".join([f"<td class='qty'>{qt}</td>" for _, qt in length_pairs])
        if include_total:
            return f"""
    <table class='matrix'>
      <tr><td class='hdr'>LEN</td>{len_cells}<td class='hdr tot'>TOT</td></tr>
      <tr><td class='hdr'>QTY</td>{qty_cells}<td class='qty tot'>{total_qty}</td></tr>
    </table>"""
        return f"""
    <table class='matrix'>
      <tr><td class='hdr'>LEN</td>{len_cells}</tr>
      <tr><td class='hdr'>QTY</td>{qty_cells}</tr>
    </table>"""

    def build_card(c):
        pairs = c["lengths"]
        chunks = [pairs[:6]]
        rest = pairs[6:]
        while rest:
            chunks.append(rest[:7])
            rest = rest[7:]

        tables = []
        total_layers = len(chunks)
        for idx, chunk in enumerate(chunks):
            is_last = idx == total_layers - 1
            tables.append(build_table(chunk, total_qty=c["total_qty"] if is_last else None, include_total=is_last))

        scale = max(1.0 - 0.3 * max(0, total_layers - 2), 0.4)
        size_pct = int(scale * 100)
        tables_html = f'<div class="tables scale-{total_layers}" style="--tbl-scale:{scale};">' + "\n".join(tables) + "</div>"
        return f"""
<div class='label-page'>
  <div class='label'>
    <div class='top'>
      <div class='top-left'>
        <div class='region-line'><span>REGION</span><span class='value'>{c["region"]}</span></div>
        <div class='bundle-line'><span>BUNDLE NO</span><span class='value'>{c["current"]}/{c["total"]}</span></div>
      </div>
      <div class='top-right'>
        <img class='qr' alt='QR' src='https://api.qrserver.com/v1/create-qr-code/?size=200x200&data={c["bundle_id"]}' width='100' height='100' />
        <div class='bundle-id'>{c["bundle_id"]}</div>
      </div>
    </div>
    <div class='so'>{c["so"]}</div>
    <div class='details'>
      <div class='k'>THICKNESS</div><div class='v'>{c["thickness"]}</div>
      <div class='k'>TYPE</div><div class='v'>{c["type"]}</div>
      <div class='k'>WIDTH</div><div class='v'>{c["width"]}</div>
      <div class='k'>TOP COLOR</div><div class='v'>{c["colour"]}</div>
    </div>
    {tables_html}
  </div>
</div>"""

    pages = "\n".join([build_card(c) for c in cards])
    css = """* { box-sizing: border-box; margin: 0; padding: 0; }
html, body { font-family: Arial, Helvetica, sans-serif; color: #111; background: #ccc; }
.label-page { width: 4in; height: 6in; background: #fff; display: flex; align-items: stretch;
  page-break-after: always; break-after: page; margin: 20px auto; box-shadow: 0 4px 16px rgba(0,0,0,0.12); }
.label { width: 100%; padding: 0.18in 0.18in 0.16in; display: flex; flex-direction: column; gap: 0.1in; }
.top { display: grid; grid-template-columns: 1fr 1.1in; gap: 0.1in; align-items: start; }
.top-left { display: flex; flex-direction: column; gap: 0.12in; }
.region-line, .bundle-line { display: flex; gap: 0.12in; align-items: baseline; font-size: 16pt; }
.region-line .value, .bundle-line .value { font-weight: 700; font-size: 17pt; }
.top-right { display: flex; flex-direction: column; align-items: center; gap: 5px; }
.qr { width: 1in; height: 1in; border: 1px solid #111; background: #fff; }
.bundle-id { font-size: 10pt; text-align: center; word-break: break-all; }
.so { font-size: 72pt; line-height: 1; text-align: center; font-weight: 400; padding: 4px 0; width: 100%; display: flex; justify-content: center; box-sizing: border-box; }
.details { display: grid; grid-template-columns: 1.3fr 1fr; row-gap: 0.08in; column-gap: 0.08in; padding: 0 4px; }
.k { font-size: 14pt; font-weight: 600; }
.v { font-size: 14pt; word-break: break-word; }
table.matrix { width: 100%; border-collapse: collapse; table-layout: auto; margin-top: auto; }
.matrix td { border: 1.5px solid #111; padding: 5px 4px; text-align: center; vertical-align: middle; }
.matrix .hdr { font-size: 11pt; font-weight: 700; background: #f5f5f5; white-space: nowrap; }
.matrix .len { font-size: 12pt; font-weight: 700; }
.matrix .qty { font-size: 13pt; font-weight: 700; }
.matrix .tot { background: #efefef; }
.tables { transform: scale(var(--tbl-scale)); transform-origin: top left; width: calc(100% / var(--tbl-scale)); }
@page { size: 4in 6in; margin: 0; }
@media print { html, body { background: #fff; }
  .label-page { width: 4in; height: 6in; margin: 0; box-shadow: none; } }"""
    return f"""<!DOCTYPE html>
<html lang='en'><head><meta charset='UTF-8'><title>Bundle Labels</title>
<style>{css}</style></head><body>{pages}</body></html>"""


def ensure_db():
    print(f"[DB] Connecting to {DB_PATH}")
    conn = sqlite3.connect(str(DB_PATH))
    cur = conn.cursor()
    cur.execute(
        "CREATE TABLE IF NOT EXISTS processed_labels ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT, "
        "work_id TEXT NOT NULL, "
        "top_color TEXT NOT NULL, "
        "created_at TEXT DEFAULT CURRENT_TIMESTAMP, "
        "UNIQUE(work_id, top_color))"
    )
    conn.commit()
    return conn


def get_existing_keys():
    conn = ensure_db()
    cur = conn.cursor()
    cur.execute("SELECT work_id, top_color FROM processed_labels")
    keys = set(cur.fetchall())
    conn.close()
    print(f"[DB] Found {len(keys)} existing entries")
    return keys


def save_keys_to_db(keys_to_save):
    if not keys_to_save:
        return
    conn = ensure_db()
    cur = conn.cursor()
    saved = 0
    for work_id, top_color in keys_to_save:
        cur.execute(
            "INSERT OR IGNORE INTO processed_labels (work_id, top_color) VALUES (?, ?)",
            (work_id, top_color),
        )
        saved += cur.rowcount
    conn.commit()
    conn.close()
    print(f"[DB] Saved {saved} new entries")


def write_to_google_sheet(bundle_df, sheet_url, creds_path):
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("[GOOGLE SHEETS] gspread or google-auth not installed")
        return False, "gspread or google-auth not installed. Run: pip install gspread google-auth"

    creds_file = Path(creds_path)
    if not creds_file.exists():
        print(f"[GOOGLE SHEETS] Credentials not found: {creds_path}")
        return False, f"Service account file not found: {creds_path}"

    try:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        creds = Credentials.from_service_account_file(str(creds_file), scopes=scopes)
        gc = gspread.authorize(creds)
        print("[GOOGLE SHEETS] Authorized, opening sheet...")

        sh = gc.open_by_url(sheet_url)
        ws = sh.sheet1

        existing_data = ws.get_all_values()
        cols = [
            "BUNDLE ID", "WORK ID", "S.O", "TYPE", "THICKNESS",
            "COLOUR", "WIDTH (m)", "REGION", "PANEL TYPE",
            "COMPOSITION", "BUNDLE PANELS", "BUNDLE SIZE",
        ]

        if len(existing_data) == 0:
            ws.append_row(cols, value_input_option="RAW")
            print("[GOOGLE SHEETS] Added header row")

        rows_to_add = []
        for _, row in bundle_df.iterrows():
            r = []
            for c in cols:
                val = row.get(c, "")
                if pd.notna(val):
                    r.append(str(val))
                else:
                    r.append("")
            rows_to_add.append(r)

        if rows_to_add:
            ws.append_rows(rows_to_add, value_input_option="RAW")
            print(f"[GOOGLE SHEETS] Appended {len(rows_to_add)} rows")

        return True, f"{len(rows_to_add)} rows written"
    except Exception as e:
        err_type = type(e).__name__
        err_msg = str(e).strip()
        if not err_msg:
            err_msg = repr(e)
        detail = f"{err_type}: {err_msg}"
        print(f"[GOOGLE SHEETS] Error: {detail}")
        if hasattr(e, 'response'):
            try:
                print(f"[GOOGLE SHEETS] Response: {e.response.text}")
                detail += f" | API response: {e.response.text[:200]}"
            except Exception:
                pass
        return False, detail


def build_excel(bundle_df, order_df, not_in_jc_df):
    wb = Workbook()
    HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    HDR_FILL  = PatternFill("solid", fgColor="1F497D")
    EXC_FILL  = PatternFill("solid", fgColor="FF0000")
    EXC_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    DATA_FONT = Font(name="Calibri", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    thin      = Side(style="thin", color="CCCCCC")
    BRD       = Border(left=thin, right=thin, top=thin, bottom=thin)
    ALT_FILL  = PatternFill("solid", fgColor="EEF2F7")
    WHT_FILL  = PatternFill("solid", fgColor="FFFFFF")

    def write_sheet(ws, df, flag_col=None):
        cols = df.columns.tolist()
        ws.row_dimensions[1].height = 28
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BRD
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            fill = ALT_FILL if ri % 2 == 0 else WHT_FILL
            for ci, col in enumerate(cols, 1):
                val = row[col]
                c = ws.cell(row=ri, column=ci, value=(val if pd.notna(val) else ""))
                c.font = DATA_FONT; c.fill = fill; c.alignment = CENTER; c.border = BRD
                if flag_col and col == flag_col and val == "**COLOUR EXCEPTION**":
                    c.fill = EXC_FILL; c.font = EXC_FONT
        for ci, col in enumerate(cols, 1):
            mx = max([len(str(col))] + [len(str(df.iloc[r][col])) for r in range(min(len(df),100))])
            ws.column_dimensions[get_column_letter(ci)].width = min(mx + 3, 45)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df)+1}"

    ws1 = wb.active
    ws1.title = "BUNDLE PLAN"
    write_sheet(ws1, bundle_df, flag_col="COLOUR")

    ws2 = wb.create_sheet("ORDER DATA")
    order_display = order_df.drop(columns=["_layer_key"], errors="ignore")
    write_sheet(ws2, order_display, flag_col="COLOUR")

    if not not_in_jc_df.empty:
        ws3 = wb.create_sheet("PENDING - NO JOB CARD")
        write_sheet(ws3, not_in_jc_df)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


with st.sidebar:
    st.header("Files")
    plan_file = st.file_uploader("1. Planning File (.xlsx)", type=["xlsx"])
    jc_file   = st.file_uploader("2. Job Card Export (.xlsx)", type=["xlsx"])
    st.markdown("---")
    st.header("Google Sheets")
    gs_url   = st.text_input("Sheet URL", value="", key="gs_url")
    gs_creds = st.text_input("Service Account JSON", value=DEFAULT_CREDS_PATH, key="gs_creds")
    st.markdown("---")
    run_btn = st.button("Generate Bundle Plan", type="primary", use_container_width=True)

if run_btn:
    if not plan_file or not jc_file:
        st.error("Please upload both files before generating.")
        st.stop()

    if "results" in st.session_state:
        del st.session_state["results"]

    print(f"\n{'='*60}")
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Starting bundle plan generation")

    with st.spinner("Processing..."):
        plan_df = pd.read_excel(plan_file)
        jc_df   = pd.read_excel(jc_file)
        print(f"[LOAD] Planning file: {len(plan_df)} rows | Job card: {len(jc_df)} rows")

        plan_df["SO_key"] = plan_df["SO NUMBER"].apply(
            lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() != "" else None)
        jc_df["SO_key"] = jc_df["Job ID"].apply(
            lambda x: str(int(float(x))) if pd.notna(x) else None)

        plan_clean = plan_df.dropna(subset=["SO_key"]).copy()
        plan_clean["COLOUR_EXTRACTED"] = plan_clean["TOP SHEET"].apply(extract_colour)
        plan_clean[["mat_ptype","mat_thick"]] = plan_clean["MATERIAL"].apply(
            lambda x: pd.Series(parse_material(x)))
        plan_clean = plan_clean.dropna(subset=["mat_ptype","mat_thick"])
        plan_clean["mat_thick"] = plan_clean["mat_thick"].astype(int)
        if "REGION" not in plan_clean.columns:
            plan_clean["REGION"] = ""
        plan_clean["REGION"] = plan_clean["REGION"].fillna("").astype(str).str.strip()

        so_counts = plan_clean.groupby("SO_key").size()
        exception_so = set(so_counts[so_counts > 1].index)
        if exception_so:
            print(f"[WARN] {len(exception_so)} SO numbers duplicated in planning file - excluding")
            st.warning(f"{len(exception_so)} SO numbers appear more than once in the planning file and were excluded.")
            exception_df = plan_clean[plan_clean["SO_key"].isin(exception_so)].copy()
            st.dataframe(exception_df, use_container_width=True)
        plan_clean = plan_clean.loc[~plan_clean["SO_key"].isin(exception_so)].copy()

        colour_map = {}
        for (so, ptype, thick), grp in plan_clean.groupby(["SO_key","mat_ptype","mat_thick"]):
            uc = [c for c in grp["COLOUR_EXTRACTED"].dropna().unique() if c not in ("", "UNKNOWN")]
            if len(uc) == 1:
                colour_map[(so, ptype, thick)] = uc[0]
            elif len(uc) > 1:
                colour_map[(so, ptype, thick)] = "**COLOUR EXCEPTION**"
            else:
                colour_map[(so, ptype, thick)] = "UNKNOWN"

        planned_keys = set(colour_map.keys())
        print(f"[PLAN] {len(planned_keys)} unique (SO, type, thickness) combos")

        region_map = {}
        for (so, ptype, thick), grp in plan_clean.groupby(["SO_key", "mat_ptype", "mat_thick"]):
            regs = [r for r in grp["REGION"].dropna().astype(str).str.strip().unique() if r]
            region_map[(so, ptype, thick)] = regs[0] if regs else ""

        jc_valid = jc_df[
            jc_df["SO_key"].isin(set(plan_clean["SO_key"])) &
            (jc_df["Structure"].str.strip().str.upper() == "PUFF PANEL")
        ].copy()
        jc_valid["Length_num"] = pd.to_numeric(jc_valid["Length"], errors="coerce")
        jc_valid = jc_valid.dropna(subset=["Length_num"])
        jc_valid["Length_num"] = jc_valid["Length_num"].astype(int)
        jc_valid[["ptype_code","thickness"]] = jc_valid["Description"].apply(
            lambda x: pd.Series(parse_description(x)))
        jc_valid = jc_valid.dropna(subset=["ptype_code","thickness"])
        jc_valid["thickness"] = jc_valid["thickness"].astype(int)
        jc_valid["Width_m"] = jc_valid["Width"] / 1000
        print(f"[JC] {len(jc_valid)} valid PUFF PANEL job card rows")

        jc_keys = set(
            (str(r["SO_key"]), r["ptype_code"], r["thickness"])
            for _, r in jc_valid.iterrows()
        )

        not_in_jc_rows = []
        for (so, ptype, thick) in planned_keys:
            if (so, ptype, thick) not in jc_keys:
                plan_row = plan_clean[
                    (plan_clean["SO_key"] == so) &
                    (plan_clean["mat_ptype"] == ptype) &
                    (plan_clean["mat_thick"] == thick)
                ].iloc[0]
                not_in_jc_rows.append({
                    "S.O": so,
                    "MATERIAL": plan_row.get("MATERIAL",""),
                    "CUSTOMER": plan_row.get("CUSTOMER",""),
                    "COLOUR": colour_map.get((so, ptype, thick),""),
                })
        not_in_jc_df = pd.DataFrame(not_in_jc_rows)
        if not not_in_jc_df.empty:
            print(f"[INFO] {len(not_in_jc_df)} planned orders have no job cards")

        order_rows = []
        for (so, ptype, thick), grp in jc_valid.groupby(["SO_key","ptype_code","thickness"]):
            if (so, ptype, thick) not in planned_keys:
                continue
            plong   = "ROOF" if ptype == "R" else "WALL"
            work_id = f"{so}-{ptype}{thick}"
            colour  = colour_map.get((so, ptype, thick), "UNKNOWN")
            width   = grp["Width_m"].iloc[0]
            len_qty = grp.groupby("Length_num")["Bal Qty"].sum().reset_index()
            len_qty = len_qty.sort_values("Length_num", ascending=False)
            row = {
                "S.O": so, "TYPE": plong, "WORK ID": work_id,
                "REGION": region_map.get((so, ptype, thick), ""), "THICKNESS": f"{thick}MM",
                "COLOUR": colour, "WIDTH": round(width, 3),
                "_layer_key": f"{thick}MM {plong} PANEL",
            }
            for i, (_, r) in enumerate(len_qty.iterrows(), start=1):
                if i > 20: break
                row[f"LENGTH-{i}"]   = int(r["Length_num"])
                row[f"QUANTITY-{i}"] = int(r["Bal Qty"])
            order_rows.append(row)

        if not order_rows:
            st.warning("No matching records found between planning file and job card export.")
            if not not_in_jc_df.empty:
                st.info(f"{len(not_in_jc_df)} planned orders have no job cards released yet.")
                st.dataframe(not_in_jc_df, use_container_width=True)
            print("[RESULT] No matching records - stopping")
            st.stop()

        order_df = pd.DataFrame(order_rows)
        print(f"[ORDER] Built {len(order_df)} work orders")

        existing_keys = get_existing_keys()
        order_df["_db_key"] = list(zip(order_df["WORK ID"], order_df["COLOUR"]))
        dup_mask = order_df["_db_key"].isin(existing_keys)
        dup_df = order_df.loc[dup_mask].drop(columns=["_db_key"]).copy()
        new_order_df = order_df.loc[~dup_mask].drop(columns=["_db_key"]).copy()

        if not dup_df.empty:
            print(f"[DEDUP] {len(dup_df)} duplicate work orders found")

        if new_order_df.empty:
            st.warning("All uploaded rows have already been planned. Nothing new to process.")
            if not dup_df.empty:
                st.dataframe(dup_df[["S.O", "WORK ID", "THICKNESS", "COLOUR", "REGION"]], use_container_width=True)
            print("[RESULT] All duplicates - stopping")
            st.stop()

        order_df = new_order_df.copy()
        keys_to_save = [(r["WORK ID"], r["COLOUR"]) for _, r in order_df.iterrows()]
        print(f"[DEDUP] {len(order_df)} new work orders to process")

        bundle_rows = []
        bundle_counter = {}

        for _, row in order_df.iterrows():
            layer_key   = row["_layer_key"]
            config      = LAYER_DATA.get(layer_key, {"normal_size": 16, "threshold": None, "large_size": None})
            normal_size = config["normal_size"]
            threshold   = config["threshold"]
            large_size  = config["large_size"]

            groups = []
            for i in range(1, 21):
                lk, qk = f"LENGTH-{i}", f"QUANTITY-{i}"
                if lk in row and pd.notna(row.get(lk)) and pd.notna(row.get(qk)):
                    groups.append(PanelGroup(length=float(row[lk]), quantity=int(row[qk])))
            if not groups:
                continue

            if threshold and large_size:
                normal_groups = [pg for pg in groups if pg.length <= threshold]
                large_groups  = [pg for pg in groups if pg.length >  threshold]
            else:
                normal_groups = groups
                large_groups  = []

            is_roof = row["TYPE"] == "ROOF"

            all_bundles = []
            if large_groups:
                all_bundles += [(b, large_size)  for b in create_bundles(large_groups,  large_size, prefer_even=is_roof)]
            if normal_groups:
                all_bundles += [(b, normal_size) for b in create_bundles(normal_groups, normal_size, prefer_even=is_roof)]

            so_val  = row["S.O"]
            ptype   = "R" if row["TYPE"] == "ROOF" else "W"
            thk_str = row["THICKNESS"].replace("MM","").strip()
            ckey    = (so_val, ptype, thk_str)

            for bundle, eff_size in all_bundles:
                seq = bundle_counter.get(ckey, 0) + 1
                bundle_counter[ckey] = seq
                bundle_id = f"{so_val}-{ptype}{thk_str}-{seq}"
                bundle_rows.append({
                    "BUNDLE ID":     bundle_id,
                    "WORK ID":       row["WORK ID"],
                    "S.O":           so_val,
                    "TYPE":          row["TYPE"],
                    "THICKNESS":     row["THICKNESS"],
                    "COLOUR":        row["COLOUR"],
                    "WIDTH (m)":     row["WIDTH"],
                    "REGION":        row["REGION"],
                    "PANEL TYPE":    layer_key,
                    "COMPOSITION":   bundle.composition,
                    "BUNDLE PANELS": bundle.total_panels,
                    "BUNDLE SIZE":   eff_size,
                })

        bundle_df = pd.DataFrame(bundle_rows)
        print(f"[BUNDLE] Generated {len(bundle_df)} bundles")

        if bundle_df.empty:
            st.warning("No bundles generated from new records.")
            print("[RESULT] No bundles - stopping")
            st.stop()

        excel_bytes = build_excel(bundle_df, order_df, not_in_jc_df)
        label_html = generate_label_html(bundle_df)
        print("[OUTPUT] Excel and labels generated")

        st.session_state["results"] = {
            "bundle_df": bundle_df,
            "order_df": order_df,
            "not_in_jc_df": not_in_jc_df,
            "dup_df": dup_df,
            "planned_keys_count": len(planned_keys),
            "excel_bytes": excel_bytes,
            "label_html": label_html,
            "keys_to_save": keys_to_save,
            "saved": False,
            "save_msg": "",
        }
        print("[DONE] Results cached in session state")

if "results" in st.session_state:
    r = st.session_state["results"]
    bundle_df = r["bundle_df"]
    order_df = r["order_df"]
    not_in_jc_df = r["not_in_jc_df"]
    dup_df = r["dup_df"]

    if not dup_df.empty:
        st.warning(f"DUPLICATE: {len(dup_df)} work order(s) already planned - skipped.")
        st.dataframe(dup_df[["S.O", "WORK ID", "THICKNESS", "COLOUR", "REGION"]], use_container_width=True)

    exceptions = order_df[order_df["COLOUR"] == "**COLOUR EXCEPTION**"]
    if not exceptions.empty:
        st.warning(f"COLOUR EXCEPTION: {len(exceptions)} WORK ID(s) have multiple colours for the same SO+material. Fix and rerun.")
        st.dataframe(exceptions[["S.O","WORK ID","THICKNESS","COLOUR"]], use_container_width=True)

    if not not_in_jc_df.empty:
        st.info(f"NOT IN JOB CARD: {len(not_in_jc_df)} planned order(s) have no job cards released yet.")
        with st.expander("View pending planned orders"):
            st.dataframe(not_in_jc_df, use_container_width=True)

    st.success(f"{len(bundle_df)} bundles across {len(order_df)} WORK IDs")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Planned Orders", r["planned_keys_count"])
    c2.metric("WORK IDs Processed", len(order_df))
    c3.metric("Total Bundles", len(bundle_df))
    c4.metric("Pending (No JC)", len(not_in_jc_df))

    st.subheader("Bundle Plan Preview")
    st.dataframe(bundle_df, use_container_width=True, height=420)

    dl_col1, dl_col2 = st.columns(2)
    with dl_col1:
        st.download_button(
            label="\u2b07 Download Bundle Plan (.xlsx)",
            data=r["excel_bytes"],
            file_name="bundle_plan_output.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with dl_col2:
        st.download_button(
            label="\u2b07 Download Print Labels (.html)",
            data=r["label_html"].encode("utf-8"),
            file_name="bundle_labels.html",
            mime="text/html",
            use_container_width=True,
        )

    st.markdown("---")

    if not r["saved"]:
        if st.button("\u2705 Confirm & Save to Database", type="primary", use_container_width=True):
            print(f"[SAVE] Confirming and saving {len(r['keys_to_save'])} entries...")
            save_keys_to_db(r["keys_to_save"])

            save_msg = "Saved to database."
            gs_url_val = st.session_state.get("gs_url", "").strip()
            gs_creds_val = st.session_state.get("gs_creds", DEFAULT_CREDS_PATH).strip()

            if gs_url_val:
                print("[SAVE] Writing to Google Sheets...")
                gs_ok, gs_detail = write_to_google_sheet(bundle_df, gs_url_val, gs_creds_val)
                if gs_ok:
                    save_msg += f" Google Sheets updated ({gs_detail})."
                else:
                    save_msg += f" Google Sheets failed: {gs_detail}"
            else:
                print("[SAVE] No Google Sheet URL configured - skipping")

            st.session_state["results"]["saved"] = True
            st.session_state["results"]["save_msg"] = save_msg
            print(f"[SAVE] {save_msg}")
            st.rerun()
    else:
        st.success(r.get("save_msg", "This plan has been confirmed and saved."))

elif not run_btn:
    st.info("Upload both files in the sidebar and click **Generate Bundle Plan**.")
