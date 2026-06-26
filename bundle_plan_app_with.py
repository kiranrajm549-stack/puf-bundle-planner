from __future__ import annotations
import streamlit as st
import pandas as pd
import re
import sqlite3
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

st.set_page_config(page_title="PUF Bundle Plan Generator", layout="wide")
st.title("PUF Bundle Plan Generator")
st.caption("Upload Planning File + Job Card Export -> Auto-generates Bundle Plan")

LAYER_DATA = {
    "30MM WALL PANEL": {"normal_size": 36, "threshold": 10000, "large_size": 10},
    "40MM WALL PANEL": {"normal_size": 26, "threshold": None, "large_size": None},
    "50MM WALL PANEL": {"normal_size": 22, "threshold": 10000, "large_size": 10},
    "60MM WALL PANEL": {"normal_size": 18, "threshold": None, "large_size": None},
    "80MM WALL PANEL": {"normal_size": 13, "threshold": None, "large_size": None},
    "100MM WALL PANEL": {"normal_size": 11, "threshold": None, "large_size": None},
    "120MM WALL PANEL": {"normal_size": 9, "threshold": None, "large_size": None},
    "150MM WALL PANEL": {"normal_size": 7, "threshold": None, "large_size": None},
    "20MM ROOF PANEL": {"normal_size": 26, "threshold": None, "large_size": None},
    "30MM ROOF PANEL": {"normal_size": 22, "threshold": 8000, "large_size": 12},
    "40MM ROOF PANEL": {"normal_size": 18, "threshold": None, "large_size": None},
    "50MM ROOF PANEL": {"normal_size": 16, "threshold": 8000, "large_size": 10},
    "60MM ROOF PANEL": {"normal_size": 14, "threshold": None, "large_size": None},
    "80MM ROOF PANEL": {"normal_size": 10, "threshold": None, "large_size": None},
    "100MM ROOF PANEL": {"normal_size": 8, "threshold": None, "large_size": None},
    "120MM ROOF PANEL": {"normal_size": 8, "threshold": None, "large_size": None},
}

@dataclass
class PanelGroup:
    length: float
    quantity: int

@dataclass
class Bundle:
    bundle_no: int
    panels: list = field(default_factory=list)
    @property
    def total_panels(self):
        return sum(pg.quantity for pg in self.panels)
    @property
    def composition(self):
        parts = []
        for i, pg in enumerate(self.panels, start=1):
            l = str(int(pg.length)) if pg.length == int(pg.length) else str(pg.length)
            parts.append(f"{pg.quantity} x L{i}({l})")
        return " + ".join(parts)

def create_bundles(panel_groups, bundle_size):
    total = sum(pg.quantity for pg in panel_groups)
    if total == 0:
        return []
    n = max(1, (total + bundle_size - 1) // bundle_size)
    sizes = [total // n + (1 if i < total % n else 0) for i in range(n)]
    sorted_groups = sorted(panel_groups, key=lambda x: x.length, reverse=True)
    pool = [[pg.length, pg.quantity] for pg in sorted_groups]
    bundles = []
    bundle_no = 1
    for target in sizes:
        current = Bundle(bundle_no=bundle_no)
        remaining_target = target
        for item in pool:
            if remaining_target <= 0:
                break
            if item[1] <= 0:
                continue
            take = min(item[1], remaining_target)
            if take > 0:
                current.panels.append(PanelGroup(length=item[0], quantity=take))
                item[1] -= take
                remaining_target -= take
        bundles.append(current)
        bundle_no += 1
    return bundles

def extract_colour(val):
    if pd.isna(val):
        return "UNKNOWN"
    s = str(val).strip()
    cleaned = re.sub(r"^[0-9]+\.?[0-9]*\s*", "", s).strip()
    return cleaned.upper() if cleaned else s.upper()

def parse_material(material_val):
    if pd.isna(material_val):
        return None, None
    d = str(material_val).strip().upper()
    ptype = "R" if "ROOF" in d else ("W" if "WALL" in d else None)
    m = re.search(r"(\d+)\s*MM", d)
    thickness = int(m.group(1)) if m else None
    return ptype, thickness

def parse_description(desc):
    if pd.isna(desc):
        return None, None
    d = str(desc).strip().upper()
    if "ROOF" in d or "TILE" in d or "SOLAR" in d:
        ptype = "R"
    elif "WALL" in d:
        ptype = "W"
    else:
        return None, None
    m = re.search(r"THICK[:\s]+(\d+)", d)
    thickness = int(m.group(1)) if m else None
    return ptype, thickness

def ensure_db():
    conn = sqlite3.connect("bundle_duplicates.db")
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS processed_labels (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            work_id TEXT NOT NULL,
            top_color TEXT NOT NULL,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(work_id, top_color)
        )
    """)
    conn.commit()
    return conn

def get_existing_keys(conn):
    cur = conn.cursor()
    cur.execute("SELECT work_id, top_color FROM processed_labels")
    return set(cur.fetchall())

def save_keys(conn, rows):
    cur = conn.cursor()
    for work_id, top_color in rows:
        cur.execute("INSERT OR IGNORE INTO processed_labels (work_id, top_color) VALUES (?, ?)", (work_id, top_color))
    conn.commit()

def generate_label_html(bundle_df: pd.DataFrame) -> str:
    return "<html><body><h3>Labels ready</h3></body></html>"

with st.sidebar:
    st.header("Files")
    planfile = st.file_uploader("1. Planning File (.xlsx)", type=["xlsx"])
    jcfile = st.file_uploader("2. Job Card Export (.xlsx)", type=["xlsx"])

runbtn = st.button("Generate Bundle Plan", type="primary", use_container_width=True)

if runbtn:
    if not planfile or not jcfile:
        st.error("Please upload both files before generating.")
        st.stop()

    with st.spinner("Processing..."):
        plandf = pd.read_excel(planfile)
        jcdf = pd.read_excel(jcfile)

        if "SO NUMBER" not in plandf.columns:
            st.error("Planning file must contain SO NUMBER.")
            st.stop()
        if "Job ID" not in jcdf.columns:
            st.error("Job card file must contain Job ID.")
            st.stop()

        plandf["SOkey"] = plandf["SO NUMBER"].apply(lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() else None)
        jcdf["SOkey"] = jcdf["Job ID"].apply(lambda x: str(int(float(x))) if pd.notna(x) and str(x).strip() else None)

        planclean = plandf.dropna(subset=["SOkey"]).copy()
        planclean["COLOUR"] = planclean["TOP SHEET"].apply(extract_colour)
        planclean[["matptype", "matthick"]] = planclean["MATERIAL"].apply(lambda x: pd.Series(parse_material(x)))
        planclean = planclean.dropna(subset=["matptype", "matthick"]).copy()
        planclean["matthick"] = planclean["matthick"].astype(int)
        if "REGION" not in planclean.columns:
            planclean["REGION"] = ""
        planclean["REGION"] = planclean["REGION"].fillna("").astype(str).str.strip()

        planok = planclean.copy()
        colourmap = {}
        regionmap = {}
        for (so, ptype, thick), grp in planok.groupby(["SOkey", "matptype", "matthick"]):
            colourmap[(so, ptype, thick)] = grp["COLOUR"].dropna().iloc[0] if len(grp["COLOUR"].dropna().unique()) == 1 else "COLOUR EXCEPTION"
            regionmap[(so, ptype, thick)] = grp["REGION"].dropna().iloc[0] if len(grp["REGION"].dropna().unique()) >= 1 else ""

        plannedkeys = set(colourmap.keys())

        jcvalid = jcdf[jcdf["SOkey"].isin(set(planok["SOkey"]))].copy()
        if "Structure" in jcvalid.columns:
            jcvalid = jcvalid[jcvalid["Structure"].astype(str).str.strip().str.upper() == "PUFF PANEL"].copy()
        if "Length" not in jcvalid.columns or "Description" not in jcvalid.columns or "Bal Qty" not in jcvalid.columns or "Width" not in jcvalid.columns:
            st.error("Job card file is missing required columns: Length, Description, Bal Qty, Width.")
            st.stop()

        jcvalid["Lengthnum"] = pd.to_numeric(jcvalid["Length"], errors="coerce")
        jcvalid = jcvalid.dropna(subset=["Lengthnum"]).copy()
        jcvalid["Lengthnum"] = jcvalid["Lengthnum"].astype(int)
        jcvalid[["ptypecode", "thickness"]] = jcvalid["Description"].apply(lambda x: pd.Series(parse_description(x)))
        jcvalid = jcvalid.dropna(subset=["ptypecode", "thickness"]).copy()
        jcvalid["thickness"] = jcvalid["thickness"].astype(int)
        jcvalid["Widthm"] = pd.to_numeric(jcvalid["Width"], errors="coerce") / 1000

        orderrows = []
        for (so, ptype, thick), grp in jcvalid.groupby(["SOkey", "ptypecode", "thickness"]):
            if (so, ptype, thick) not in plannedkeys:
                continue
            row = {
                "S.O": so,
                "TYPE": "ROOF" if ptype == "R" else "WALL",
                "WORK ID": f"{so}-{ptype}{thick}",
                "REGION": regionmap.get((so, ptype, thick), ""),
                "THICKNESS": f"{thick}MM",
                "COLOUR": colourmap.get((so, ptype, thick), "UNKNOWN"),
                "WIDTH (m)": round(float(grp["Widthm"].iloc[0]), 3) if pd.notna(grp["Widthm"].iloc[0]) else None,
                "PANEL TYPE": f"{thick}MM {'ROOF' if ptype == 'R' else 'WALL'} PANEL",
            }
            lenqty = grp.groupby("Lengthnum", as_index=False)["Bal Qty"].sum().sort_values("Lengthnum", ascending=False)
            for i, (_, r) in enumerate(lenqty.iterrows(), start=1):
                if i > 20:
                    break
                row[f"LENGTH-{i}"] = int(r["Lengthnum"])
                row[f"QUANTITY-{i}"] = int(r["Bal Qty"])
            orderrows.append(row)

        if not orderrows:
            st.warning("No matching records found between planning file and job card export.")
            st.stop()

        orderdf = pd.DataFrame(orderrows)
        conn = ensure_db()
        existing = get_existing_keys(conn)
        orderdf["DB_KEY"] = list(zip(orderdf["WORK ID"], orderdf["COLOUR"]))
        dup_mask = orderdf["DB_KEY"].isin(existing)
        dup_df = orderdf.loc[dup_mask].copy()
        new_orderdf = orderdf.loc[~dup_mask].copy()

        if not dup_df.empty:
            st.warning(f"Duplicate work id + top color found: {len(dup_df)}")
            st.dataframe(dup_df[["S.O", "WORK ID", "THICKNESS", "COLOUR", "REGION"]], use_container_width=True)

        if new_orderdf.empty:
            conn.close()
            st.warning("All uploaded rows are duplicates. Nothing new to process.")
            st.stop()

        orderdf = new_orderdf.drop(columns=["DB_KEY"]).copy()

        bundlerows = []
        bundlecounter = {}
        for _, row in orderdf.iterrows():
            layerkey = row["PANEL TYPE"]
            config = LAYER_DATA.get(layerkey, {"normal_size": 16, "threshold": None, "large_size": None})
            groups = []
            for i in range(1, 21):
                lk, qk = f"LENGTH-{i}", f"QUANTITY-{i}"
                if lk in row and pd.notna(row.get(lk)) and pd.notna(row.get(qk)):
                    groups.append(PanelGroup(length=float(row[lk]), quantity=int(row[qk])))
            if not groups:
                continue
            normalsize = config["normal_size"]
            threshold = config["threshold"]
            largesize = config["large_size"]
            if threshold and largesize:
                normalgroups = [pg for pg in groups if pg.length <= threshold]
                largegroups = [pg for pg in groups if pg.length > threshold]
            else:
                normalgroups = groups
                largegroups = []
            allbundles = []
            if normalgroups:
                allbundles.extend([(b, normalsize) for b in create_bundles(normalgroups, normalsize)])
            if largegroups:
                allbundles.extend([(b, largesize) for b in create_bundles(largegroups, largesize)])
            soval = row["S.O"]
            ptype = "R" if row["TYPE"] == "ROOF" else "W"
            thkstr = row["THICKNESS"].replace("MM", "").strip()
            ckey = (soval, ptype, thkstr)
            seq = bundlecounter.get(ckey, 0)
            for bundle, effsize in allbundles:
                seq += 1
                bundlecounter[ckey] = seq
                bundleid = f"{soval}-{ptype}{thkstr}-{seq:02d}"
                bundlerows.append({
                    "BUNDLE ID": bundleid,
                    "WORK ID": row["WORK ID"],
                    "S.O": soval,
                    "TYPE": row["TYPE"],
                    "THICKNESS": row["THICKNESS"],
                    "COLOUR": row["COLOUR"],
                    "WIDTH (m)": row["WIDTH (m)"],
                    "REGION": row["REGION"],
                    "PANEL TYPE": layerkey,
                    "COMPOSITION": bundle.composition,
                    "BUNDLE PANELS": bundle.total_panels,
                    "BUNDLE SIZE": effsize,
                })

        bundledf = pd.DataFrame(bundlerows)
        save_keys(conn, [(r["WORK ID"], r["COLOUR"]) for _, r in orderdf.iterrows()])
        conn.close()

        if bundledf.empty:
            st.warning("No bundles were generated from the new records.")
            st.stop()

        st.success(f"{len(bundledf)} bundles across {len(orderdf)} WORK IDs")
        st.dataframe(bundledf, use_container_width=True, height=420)

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "BUNDLE PLAN"
        for cidx, col in enumerate(bundledf.columns, start=1):
            ws1.cell(1, cidx, col)
            for ridx, val in enumerate(bundledf[col], start=2):
                ws1.cell(ridx, cidx, val)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        st.download_button("Download Bundle Plan .xlsx", data=buf, file_name="bundleplanoutput.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        st.download_button("Download Print Labels .html", data=generate_label_html(bundledf).encode("utf-8"), file_name="bundlelabels.html", mime="text/html", use_container_width=True)
else:
    st.info("Upload both files in the sidebar and click Generate Bundle Plan.")
