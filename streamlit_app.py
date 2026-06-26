"""
app.py  –  PUF Panel Bundle Planner  (Streamlit)
Run:  streamlit run app.py
"""

import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bundle_logic import (
    bundles_to_dataframe,
    build_summary,
    detect_order_columns,
)

st.set_page_config(page_title="PUF Bundle Planner", page_icon="📦", layout="wide")

ORDER_SHEET_ALIASES = ["ORDER DATA", "ORDERDATA", "ORDER", "ORDERS", "Sheet1"]
LAYER_SHEET_ALIASES = ["LAYER DATA", "LAYERDATA", "LAYER", "LAYERS", "Sheet2"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL    = PatternFill("solid", fgColor="D9E1F2")
BORDER_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _find_sheet(xls, aliases):
    available = {s.upper(): s for s in xls.sheet_names}
    for alias in aliases:
        if alias.upper() in available:
            return available[alias.upper()]
    return None


def _read_header_row(xls, sheet):
    KNOWN = {"S.O", "SO", "TYPE", "THICKNESS", "L1", "Q1",
             "WORK ID", "WORK_ID", "PANEL TYPE"}
    raw = xls.parse(sheet, header=None)
    for i, row in raw.iterrows():
        row_vals = {str(v).strip().upper() for v in row if pd.notna(v)}
        if row_vals & KNOWN:
            df = xls.parse(sheet, header=i)
            df.columns = [str(c).strip() for c in df.columns]
            return df.dropna(how="all").reset_index(drop=True)
    df = xls.parse(sheet)
    df.columns = [str(c).strip() for c in df.columns]
    return df.dropna(how="all").reset_index(drop=True)


def _style_sheet(ws, df, col_widths=None):
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER_THIN

    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = ALT_FILL if r_idx % 2 == 0 else PatternFill()
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = fill
            cell.alignment = CENTER
            cell.border = BORDER_THIN

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = (col_widths or {}).get(col_name, 14)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def build_excel(bundle_df, summary_df):
    wb = Workbook()

    ws_plan = wb.active
    ws_plan.title = "BUNDLE PLAN"
    plan_cols = [c for c in [
        "BUNDLE_ID", "WORK_ID", "S.O", "TYPE", "REGION",
        "THICKNESS", "COMPOSITION", "BUNDLE_PANELS", "BUNDLE_SIZE", "COLOUR", "WIDTH"
    ] if c in bundle_df.columns]
    _style_sheet(ws_plan, bundle_df[plan_cols], {
        "BUNDLE_ID": 20, "WORK_ID": 16, "S.O": 14,
        "TYPE": 10, "REGION": 10, "THICKNESS": 12,
        "COMPOSITION": 40, "BUNDLE_PANELS": 14,
        "BUNDLE_SIZE": 13, "COLOUR": 10,
    })

    ws_sum = wb.create_sheet("SUMMARY")
    sum_cols = [c for c in [
        "WORK_ID", "S.O", "TYPE", "THICKNESS",
        "BUNDLE_SIZE", "TOTAL_BUNDLES", "PARTIAL_BUNDLES", "TOTAL_PANELS"
    ] if c in summary_df.columns]
    _style_sheet(ws_sum, summary_df[sum_cols], {
        "WORK_ID": 16, "S.O": 14, "TYPE": 10, "THICKNESS": 12,
        "BUNDLE_SIZE": 13, "TOTAL_BUNDLES": 14,
        "PARTIAL_BUNDLES": 15, "TOTAL_PANELS": 13,
    })

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── UI ────────────────────────────────────────────────────────────────────────
st.title("📦 PUF Panel Bundle Planner")
st.caption("Upload your Excel file with **ORDER DATA** and **LAYER DATA** sheets.")

uploaded = st.file_uploader("Choose Excel file (.xlsx)", type=["xlsx", "xls"])

if not uploaded:
    st.info("👆 Upload an Excel file to get started.")
    st.stop()

try:
    xls = pd.ExcelFile(uploaded)
except Exception as e:
    st.error(f"Could not read file: {e}")
    st.stop()

order_sheet = _find_sheet(xls, ORDER_SHEET_ALIASES)
layer_sheet = _find_sheet(xls, LAYER_SHEET_ALIASES)

if not order_sheet:
    st.error(f"❌ ORDER DATA sheet not found. Available: {', '.join(xls.sheet_names)}")
    st.stop()
if not layer_sheet:
    st.warning("⚠️ LAYER DATA sheet not found – using default bundle size of 16 panels.")

orders_df = _read_header_row(xls, order_sheet)
layer_df  = _read_header_row(xls, layer_sheet) if layer_sheet else pd.DataFrame()

with st.expander("🔍 Column detection preview", expanded=False):
    cols_map = detect_order_columns(list(orders_df.columns))
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Detected columns:**")
        for k, v in cols_map.items():
            st.write(f"{'✅' if v else '➖'} `{k}` → `{v}`")
    with c2:
        st.markdown("**Raw columns in file:**")
        st.write(list(orders_df.columns))

with st.expander("📋 Raw order data preview", expanded=False):
    st.dataframe(orders_df.head(20), use_container_width=True)

with st.spinner("Calculating bundles…"):
    try:
        bundle_df  = bundles_to_dataframe(orders_df, layer_df)
        summary_df = build_summary(bundle_df)
    except Exception as e:
        st.error(f"Bundle calculation error: {e}")
        st.stop()

if bundle_df.empty:
    st.warning("No bundles generated. Check ORDER DATA has valid panel rows.")
    st.stop()

# Metrics — uses BUNDLE_PANELS (not QUANTITY)
m1, m2, m3, m4 = st.columns(4)
m1.metric("Total Orders",   orders_df.shape[0])
m2.metric("Total Bundles",  bundle_df["BUNDLE_ID"].nunique())
m3.metric("Total Panels",   int(bundle_df["BUNDLE_PANELS"].sum()))
partial_count = int((summary_df["PARTIAL_BUNDLES"] > 0).sum()) if "PARTIAL_BUNDLES" in summary_df.columns else 0
m4.metric("Orders w/ Partial Bundle", partial_count)

st.divider()

tab_plan, tab_sum = st.tabs(["📦 Bundle Plan", "📊 Summary"])

with tab_plan:
    plan_cols = [c for c in [
        "BUNDLE_ID", "WORK_ID", "S.O", "TYPE", "REGION",
        "THICKNESS", "COMPOSITION", "BUNDLE_PANELS", "BUNDLE_SIZE", "COLOUR", "WIDTH"
    ] if c in bundle_df.columns]

    fc1, fc2, fc3 = st.columns(3)
    sel_type = fc1.selectbox("Filter TYPE",      ["All"] + sorted(bundle_df["TYPE"].dropna().unique().tolist()))
    sel_thk  = fc2.selectbox("Filter THICKNESS", ["All"] + sorted(bundle_df["THICKNESS"].dropna().unique().tolist()))
    sel_so   = fc3.selectbox("Filter S.O",       ["All"] + sorted(bundle_df["S.O"].dropna().unique().tolist()))

    filtered = bundle_df.copy()
    if sel_type != "All": filtered = filtered[filtered["TYPE"] == sel_type]
    if sel_thk  != "All": filtered = filtered[filtered["THICKNESS"] == sel_thk]
    if sel_so   != "All": filtered = filtered[filtered["S.O"] == sel_so]

    st.dataframe(filtered[plan_cols], use_container_width=True, height=420)

with tab_sum:
    sum_cols = [c for c in [
        "WORK_ID", "S.O", "TYPE", "THICKNESS",
        "BUNDLE_SIZE", "TOTAL_BUNDLES", "PARTIAL_BUNDLES", "TOTAL_PANELS"
    ] if c in summary_df.columns]
    st.dataframe(summary_df[sum_cols], use_container_width=True, height=420)

st.divider()

excel_bytes = build_excel(bundle_df, summary_df)
st.download_button(
    label="⬇️  Download Bundle Plan (.xlsx)",
    data=excel_bytes,
    file_name="puf_bundle_plan.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
    type="primary",
)